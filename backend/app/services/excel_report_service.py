import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.models import Operation, Measurement, AlarmEvent

def generate_operation_excel_report(db: Session, operation: Operation) -> io.BytesIO:
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4A5568")
    bold_font = Font(name="Calibri", size=11, bold=True)
    border_thin = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    # ----------------------------------------------------
    # SHEET 1: Resumen de Operación
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Resumen de Operación"

    ws1["A1"] = "SISTEMA WEB DE MONITOREO DE CONDICIÓN DE BOMBAS DE TRANSFERENCIA"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Reporte Consolidado de Operación - {operation.codigo_operacion}"
    ws1["A2"].font = subtitle_font

    ws1.append([]) # Blank row

    # Operation Metadata
    tanks_str = ", ".join([t.codigo for t in operation.tanks]) if operation.tanks else "N/A"
    pumps_str = ", ".join([f"{p.codigo} ({p.nombre})" for p in operation.pumps]) if operation.pumps else "N/A"

    info_rows = [
        ("Código de Operación:", operation.codigo_operacion, "Fecha:", str(operation.fecha)),
        ("Buque:", operation.buque.nombre if operation.buque else "N/A", "Empresa Buque:", operation.buque.empresa if operation.buque else "N/A"),
        ("Producto Transferido:", operation.producto.nombre if operation.producto else "N/A", "Estado Operación:", operation.estado),
        ("Tanques de Origen:", tanks_str, "Bombas Asignadas:", pumps_str),
        ("Hora Inicio:", str(operation.hora_inicio), "Hora Finalización:", str(operation.hora_fin or "En Proceso")),
        ("Responsable Operativo:", operation.responsable.full_name if operation.responsable else "N/A", "Observaciones:", operation.observaciones or "Ninguna")
    ]

    for row in info_rows:
        ws1.append([row[0], row[1], "", row[2], row[3]])
        curr_row = ws1.max_row
        ws1.cell(row=curr_row, column=1).font = bold_font
        ws1.cell(row=curr_row, column=4).font = bold_font

    # ----------------------------------------------------
    # SHEET 2: Mediciones de Campo
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Mediciones de Campo")
    
    headers_ws2 = [
        "#", "Fecha", "Hora", "Bomba Código", "Bomba Nombre", 
        "Presión Succión (inHg)", "Presión Descarga (psi)", 
        "Temp. Motor (°C)", "Temp. Bomba (°C)", "Corriente (A)", "Estado Alarma", 
        "Técnico Inspector", "Observaciones"
    ]
    ws2.append(headers_ws2)
    for col_num in range(1, len(headers_ws2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    measurements = db.query(Measurement).filter(
        Measurement.operation_id == operation.id
    ).order_by(Measurement.datetime_registro.asc()).all()

    for idx, m in enumerate(measurements, 1):
        is_alarm = m.temperatura_c > 80.0 or m.corriente_a > 45.0
        status_str = "ALARMA" if is_alarm else "NORMAL"
        if m.is_corrected:
            status_str += " (Corregido)"

        row_vals = [
            idx,
            str(m.fecha_registro),
            str(m.hora_registro),
            m.bomba.codigo if m.bomba else "",
            m.bomba.nombre if m.bomba else "",
            m.presion_succion_inhg if m.presion_succion_inhg is not None else "",
            m.presion_descarga_psi,
            m.temperatura_c,
            m.temperatura_bomba_c if m.temperatura_bomba_c is not None else "",
            m.corriente_a,
            status_str,
            m.registrado_por.full_name if m.registrado_por else "",
            m.observaciones or ""
        ]
        ws2.append(row_vals)
        curr_row = ws2.max_row
        
        # Format row borders and alignments
        for col_num in range(1, len(row_vals) + 1):
            cell = ws2.cell(row=curr_row, column=col_num)
            cell.border = border_thin
            if col_num in [1, 2, 3, 4, 10]:
                cell.alignment = Alignment(horizontal="center")
            elif col_num in [6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="right")

    # ----------------------------------------------------
    # SHEET 3: Eventos de Alarma
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Eventos de Alarma")
    headers_ws3 = ["#", "Fecha y Hora", "Bomba", "Tipo Alarma", "Nivel", "Valor Registrado", "Límite Umbral", "Mensaje"]
    ws3.append(headers_ws3)
    for col_num in range(1, len(headers_ws3) + 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alarms = db.query(AlarmEvent).filter(
        AlarmEvent.operacion_id == operation.id
    ).order_by(AlarmEvent.fecha_hora.asc()).all()

    for idx, a in enumerate(alarms, 1):
        ws3.append([
            idx,
            a.fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if a.fecha_hora else "",
            a.bomba.codigo if a.bomba else "",
            a.tipo_alarma,
            a.nivel,
            a.valor_registrado,
            a.limite_umbral,
            a.mensaje
        ])

    # Auto-adjust column widths and apply borders for all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    cell.border = border_thin
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
